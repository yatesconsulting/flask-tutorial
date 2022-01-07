from flask import (
    Blueprint, flash, g, redirect, render_template, request, url_for, send_file
    
)
from werkzeug.exceptions import abort

from flaskr.auth import login_required
from flaskr.db import get_db
from flask.helpers import read_image, make_response
from openpyxl import Workbook, load_workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.worksheet.datavalidation import DataValidation

import io
from datetime import datetime
import re

# from . import mssql_db as dbsource
from . import pyodbc_db as dbsource

bp = Blueprint('inventory', __name__, url_prefix='/inv')

@bp.route('/')
def index():
    """Show the menu"""
    return render_template('inventory/index.html')

def _formfromdb(db, sql, params=()):
    try:
        ans = []
        if params:
            rows = db.execute(sql, params) #.fetchall()
        else:
            rows = db.execute(sql) #.fetchall()
        return list(rows)
    except Exception as e:
        return ['_formfromdb fail', e, sql, params]

def _insertdetailrecord(dbf, values):
    # https://stackoverflow.com/questions/38878841/how-to-insert-data-into-sqlite-with-pythons-dictionary
    # dbf = get_db() # forms
    try:
        cur = dbf.cursor()
        # invid = id of form
        columns = ', '.join(values.keys())
        placeholders = ', '.join('?' * len(values))
        sql = 'INSERT INTO invdelformdetails ({}) VALUES ({})'.format(columns, placeholders)
        cur.execute(sql, tuple(values.values()))
        dbf.commit()
        cur.close # ? good idea?
        return True
    except Exception as e:
        return [e, sql]

def _deletedetailrecord(dbf, id):
    try:
        cur = dbf.cursor()
        sql = 'delete from invdelformdetails where id=?'
        cur.execute(sql, (id,)) # tuple(id) fails here, not sure why
        dbf.commit()
        cur.close # ? good idea?
        return True
    except Exception as e:
        return [e, sql]

def _updateheaderrecord(db, f):
    # put it exactly as sent in f, blank out any fields that are blank
    if 'id' not in f:
        return False
    try:
        id =  f['id']
        del f['id']
        (keys,values) = zip(*f.items())
        updatevals = list(values)
        updateme = ['{} = ?'.format(key) for key in keys]
        cur = db.cursor()
        if updateme:
            sql = "update invdelform set {}, updated=CURRENT_TIMESTAMP where id = ?;".format(', '.join(updateme))
            updatevals.append(id)
            cur.execute(sql, (updatevals))
            db.commit()
            cur.close # ?
            return True
        else:
            return ['Nothing to do with id {}'.format(id)]
    except Exception as e:
        return [e, sql]

def _updatedetailrecord(db, f):
    # put it exactly as sent in f, blank out any fields that are blank
    if 'id' not in f:
        return [False, 'id not in update dictionary']
    cur = db.cursor()
    id =  f['id']
    tag = "" if not 'tag' in f else f['tag']
    description = "" if not 'description' in f else f['description']
    delcode = "" if not 'delcode' in f else f['delcode']
    itinitials = "" if not 'itinitials' in f else f['itinitials']
    sql = "update invdelformdetails set tag = ?, description=?, delcode=?, itinitials=?, updated=CURRENT_TIMESTAMP where id = ?;"
    try:
        cur.execute(sql, (tag, description, delcode, itinitials, id))
        db.commit()
        cur.close # ?
        return True
    except Exception as e:
        return [e, sql]

def _getheaderinfo(db, delformid):
    if not delformid:
        return []
    try:
        sql = "select *,(select count(distinct tag) from invdelformdetails where invid=F.id) as tagcount from invdelform F where id = ?;"
        ans = _formfromdb(db, sql, tuple(delformid))
        return ans
    except Exception as e:
        return ['_getheaderinfo failed', e, sql]

def _getdetails(db, sheetid):
    ans = _formfromdb(db, 'select * from invdelformdetails where invid = ?;',tuple(sheetid))
  # , dateitcleared, updated
    # id, invid, tag, description , delcode, itinitials
    return ans

def _getdetailsfromtags(db, tags):
    # really get this from some sql similar to...
    # select  tagnumber, Description from vw_web_FA_base where tagnumber in (?)
    sql = "SELECT M.TagNumber as tag, D.description FROM RISDBase.dbo.tblFAAssetDetail D join RISDBase.dbo.tblFAAssetMaster M on M.ID = D.AssetMasterID where M.tagnumber in ({})".format(",".join("?"*len(tags)) )
    ans = db.execute_s(sql, tags)
    return ans
    # return  [{'TagNumber':'57476','Description':'MS SURFACE PRO 4 M 128GB 4GB EDU BD Mfg#: TZ5-00001 Contract: MARKET Comes with keyboard type cas'},{'TagNumber':'57368','Description':'UAG SURFACE PRO 4 CASE - BLACK'}]

# def _inserttagsintodeletionsheet(tagwithdescription):
#     # send a complete list with desc
#     for tag in tagwithdescription:
#         t = tag['TagNumber']
#         d = tag['Description']


@bp.route('/del', methods=["GET", "POST"])
@login_required
def deletions():
    id, scanned,submit, show = [None] * 4

    dbv = dbsource.MSSQL_DB_Conn() # visions info
    dbf = get_db() # forms
    username = g.user['username'][:2].upper()
    initials = username.upper()

    # return render_template('inventory/index.html',rows= ['blah',request.values])
    # return render_template('inventory/index.html',rows= ['blah',request.args])
    # return render_template('inventory/index.html',rows= ['blah',request.stream])

    # blank /inv/del or /inv/del?id=1, maybe direct links to a couple options

    # normal POST request submits can come in a few ways
    # submit buttons Download or Add end with id, those input fields will be name#
    # submit from scanning will be id=#,scanned=#,scanned=#

    # process form data into some starter variables
    submit = id = show = scanned = ""
    if request.method == 'GET':
        id = request.args.get('id')
        show = request.args.get("show")
        # return render_template('inventory/index.html',rows= ['186blah', show])
    elif request.method == 'POST':
        scanned = request.form.getlist("scanned")
        # show = request.form.get("show")
        submit = request.form.get("submit")
        id = request.form.get("id") or request.form.get("delformid")
        if not id:
            id = submit.split(" ")[-1]
        if id and not id.isdigit():
            id = "" # do not make this an int, as tuple will fail on it later

    # return render_template('inventory/index.html',rows=[{'203short circuit submit=':submit, 'id':id, 'show':show, 'scanned':scanned}])

    # now do any database operations or redirections based on inputs
    if re.match("^(Update header)", submit):
        rec = {}
        # updateformandcontinuewith, id = submit.rsplit(" ")
        rec['id'] = id # = submit.split(" ")[-1]
        rec['schooldeleting'] = request.form.get("sdel")
        rec['workorder'] =  request.form.get("wo")
        rec['notes'] = request.form.get("notes")
        rec['username'] = request.form.get("creator")
        ans = _updateheaderrecord(dbf, rec)
        if ans != True:
            flash("Error, that update failed")
        submit = "Edit {}".format(id)
        # return render_template('inventory/index.html',rows=['submit fixed?', submit, rec, ans])
    elif re.match("^(Add New Detail line to Form)", submit):
        rec = {}
        rec['invid'] = id
        rec['tag'] = request.form.get("newtag")
        rec['description'] = request.form.get("newdescription")
        rec['delcode'] = request.form.get("newdelcode")
        rec['itinitials'] =  request.form.get("newitinitials")
        ans = _insertdetailrecord(dbf, rec)
        if ans != True:
            flash("Error inserting detail record {}".format(ans))
        submit = "Edit {}".format(id)
    elif re.match("^(Update detail)", submit):
        detailid = submit.split(" ")[-1]
        rec = {}
        # updateformandcontinuewith, id = submit.rsplit(" ")
        rec['id'] = detailid # = submit.split(" ")[-1]
        rec['tag'] = request.form.get("tag{}".format(detailid))
        rec['description'] = request.form.get("description{}".format(detailid))
        rec['delcode'] = request.form.get("delcode{}".format(detailid))
        rec['itinitials'] =  request.form.get("itinitials{}".format(detailid))
        ans = _updatedetailrecord(dbf, rec)
        if ans != True:
            flash("Error updating record {}".format(ans))
        submit = "Edit {}".format(id)
        # return render_template('inventory/index.html',rows=[{'submitN':submit, 'detailid':detailid,'id':id}])
    elif re.match("^(Delete Detail)", submit):
        detailid = submit.split(" ")[-1]
        ans = _deletedetailrecord(dbf, detailid)
        if ans != True:
            flash("Error deleting record {}".format(ans))
        submit = "Edit {}".format(id)
        # return render_template('inventory/index.html',rows=[{'submitM':submit, 'detailid':detailid,'id':id}])
    elif re.match("^(Flag inactive|Delete entire form)", submit):
        rec = {}
        rec['id'] = id
        if "Flag inactive" in submit:
            rec['active'] = 0
        elif "Delete entire form" in submit:
            rec['active'] = 2
        ans = _updateheaderrecord(dbf, rec)
        if ans != True:
            flash("Error, that update failed {}".format(ans))
        submit = ""
        id = "" # refresh to entry page
        # return render_template('inventory/index.html',rows=['259submit fixed?', submit, rec, ans])
    elif re.match("^(Create & Start Scanning)", submit):
        # take top of form values and stick them in InvDelForms
        # then launch into adding scans with the ID
        notes = request.form.get("notes")
        sdel = request.form.get("sdel")
        wo = request.form.get("wo")
        creator = request.form.get("creator")
        cur = dbf.cursor() # had to get cursor to get .lastrowid
        id = cur.execute("insert into invdelform (schooldeleting, workorder, notes, username) VALUES (?,?,?,?)", (sdel, wo, notes, creator)).lastrowid
        dbf.commit()
        cur.close()
        submit = "Smartphone scan to {}".format(id)
    elif re.match("^(Submit Tags)", submit) and not scanned:
        submit = "Edit {}".format(id)
    elif re.match("^(Submit Tags)", submit) and scanned:
        added = []
        alreadyadded = []
        rows = []
        form = []
        if '\r\n' in scanned[0]:
            scanned = scanned[0].split('\r\n')
            if '' in scanned:
                scanned.remove('')
        # return render_template('inventory/index.html',rows=[{'286 submit=':submit, 'id':id, 'show':show, 'scanned':scanned}])

        details = _getdetailsfromtags(dbv, scanned)

        for d in details:
            if d['tag'] not in details:
                # add it to the form
                d['invid'] = id
                d['delcode'] = 'o' # default a better way?
                d['itinitials'] = initials # default a better way--kisok mode TODO
                # values = dict of all values for 1 row to insert into DB
                _insertdetailrecord(dbf, d)
                added.append(d)
            else:
                alreadyadded.append(d)
        if added:
            rows.append('These were added to Del form {}'.format(id))
            rows.append(added)
        if alreadyadded:
            rows.append('These were skipped, as they are already in this Del form {}'.format(id))
            rows.append(alreadyadded)
        if added:
            flash("{} new scanned IDs added to form {}".format(len(added), id))
        if alreadyadded:
            flash("{} Duplicate scans ignored".format(len(alreadyadded)))

        submit = "Edit {}".format(id)
        # return render_template('inventory/index.html',rows=rows, form=form )
    
    # return render_template('inventory/index.html',rows=[{'287short circuit2 submit=':submit, 'id':id, 'show':show, 'scanned':scanned}])

    # now send them to the correct output path
    # new form since nothing good was selected
    if not id and not submit:
        # show summaries for existing items, and allow add new delform
        form = []
        color = ""
        formtitle = ""
        sql = ""
        myform = ""
        if show == "inactive":
            sql = "select *,(select count(distinct tag) from invdelformdetails where invid=F.id) as tagcount from invdelform F where F.active = 0 order by updated desc;"
            color = "Pink"
            formtitle = "Inactive Deletion Forms"

        elif show == "deleted":
            sql = "select *,(select count(distinct tag) from invdelformdetails where invid=F.id) as tagcount from invdelform F where F.active = 2 order by updated desc;"
            color = "YellowGreen"
            formtitle = "Deleted Deletion Forms"

        else:
            sql = "select *,(select count(distinct tag) from invdelformdetails where invid=F.id) as tagcount from invdelform F where F.active = 1 order by updated desc;"
            color = "AliceBlue"
            formtitle = "Active Deletion Forms"
            myform = '<p style="text-align: right;"><a href="?show=inactive">View Inactive Forms</a></p>'
        rows = _formfromdb(dbf, sql)
        # for r in/ rows:
        # return render_template('inventory/index.html',rows=[{'317 submit=':submit, 'rows':rows, 'sql':sql}])

        for f in rows:
            i = f['id']
            cnt = f['tagcount']
            s = "s" if cnt != 1 else ""
            myform += '<div style="border-radius: 10px; border:2px solid black; background: {}; padding:10px;"><h4>Deletion form number {} - contains {} tag{}</h4>'.format(color, i, cnt, s)

            v = "" if f['notes'] == None else f['notes']
            if v:
                myform += '\nPersonal notes: {}<br />'.format(v)

            v = "" if f['schooldeleting'] == None else f['schooldeleting']
            if v:
                myform += '\nSchool Unit Deleteing Items: {}<br />'.format(v)

            v = "" if f['workorder'] == None else f['workorder']
            if v:
                myform += '\nWork Order Num: {}<br />'.format(v)

            v = "" if f['username'] == None else f['username']
            if v:
                myform += '\nOwner username: {}<br />'.format(v)

            v = "" if f['updated'] == None else f['updated']
            if v:
                myform += '\nUpdated: {}<br />'.format(v)


            # myform += '<input type="submit" name="submit" value="View {}"> '.format(i)
            # myform += '\n<input type="submit" name="submit" value="Download {}"> '.format(i)
            myform += '<br />\n<input type="submit" name="submit" value="Edit {}"> '.format(i)
            myform += '\n<input type="submit" name="submit" value="Smartphone scan to {}"> '.format(i)
            myform += '\n<input type="submit" name="submit" value="Barcode scan or 10key to {}"> '.format(i)
            myform += '\n</div><hr />'
            form.append(myform)
            myform = ""
        if show:
            if not form:
                flash("no {} forms found".format(show))
                return redirect(url_for('.deletions'))
        else:
            form.append('<h3>Make a new Deletion form and start scanning</h3>')
            form.append('<label for="notes">Personal notes:</label><input type="text" id="notes" name="notes" ><br />')
            form.append('<label for="sdel">School Unit Deleteing Items:</label> <input type="text" id="sdel" name="sdel" value=""><br />')
            form.append('<label for="wo">Work Order Num:</label> <input type="text" id="wo" name="wo"><br />')
            form.append('<label for="creator">* Username of form owner:</label> <input type="text" id="creator" name="creator" value="{}" required><br />'.format(g.user['username']))
            form.append('<input type="submit" name="submit" id="submit" value="Create & Start Scanning">')

        return render_template('inventory/index.html', form=form, formtitle=formtitle, pgtitle="Deletions")
    elif re.match("^(Add)", submit):

        # return render_template('inventory/index.html',rows=[':158b', rec['id'],  r,  id, schooldeleting, workorder, notes, creator ] )
        return render_template('inventory/scan.html', pgtitle = 'Adding to deletion form {}'.format(id), id=id)
    elif re.match("^(Download)", submit):
        headerinfo = _getheaderinfo(dbf, id)[0]
        fulldetails = _getdetails(dbf, id)
        tags =  sorted(list(set([t['tag'] for t in fulldetails])))
        wb = load_workbook(filename = '/var/www/flaskr/flaskr/static/FARetirementBLANK.xlsx')
        ws = wb.active

        # this header info will be copied to new sheets, so just do it once
        if 'schooldeleting' in headerinfo and headerinfo['schooldeleting']:
            ws['A13'].value = headerinfo['schooldeleting']
        if 'workorder' in headerinfo and headerinfo['workorder']:
            ws['I19'].value = headerinfo['workorder']
        ws['I13'] = datetime.today().strftime('%m / %d / %Y')
        # img = openpyxl.drawing.image.Image('roycesig.jpg')
        # img.anchor = 'A1'
        # ws.add_image(img)

        line = 25
        firsttag = ""
        lasttag = ""
        firstsheetname = "" # for forumla use on sheets>1
        # {'TagNumber':'57476','Description':'3Y TABLET REPAIR W/ADH $801-$900DOP'}
        for tag in tags:
            mylist = []
            mylist = [item for item in fulldetails if item["tag"] == tag]
            # return render_template('inventory/index.html',rows = ['blue', tag, mylist, fulldetails])
            dvals = []
            dvals = sorted(set([t['description'] for t in mylist]))

            if line == 25:
                firsttag = tag
            if (line > 45):
                ws.title = "{}-{}".format(firsttag, lasttag)
                if not firstsheetname:
                    firstsheetname = ws.title
                line = 25
                firsttag = tag
                ws2 = wb.copy_worksheet(ws)
                ws = ws2
                for t in range(25,46):
                    for r in (['B', 'C', 'I', 'K', 'L', 'M']):
                        ws['{}{}'.format(r, t)].value = ""
                ws['A13'] = "='{}'!A13".format(firstsheetname)
                ws['I13'] = "='{}'!I13".format(firstsheetname)
                ws['I19'] = "='{}'!I19".format(firstsheetname)

            t = ws['B{}'.format(line)]
            lasttag = tag
            t.value = tag
            t.alignment=Alignment(horizontal='center', vertical='center', shrink_to_fit=True)
            d = ws['C{}'.format(line)]
            # insert all the values here, if more than one, then make it DV, otherwise just one

            if dvals and len(dvals) > 1:
                d.value = dvals[0]
                dvallist = ','.join([re.sub('[,\'\"]', ' ', d) for d in dvals])
                # return render_template('inventory/index.html',rows = [tag, dvallist] )
                dv = DataValidation(type="list", formula1='"{}"'.format(dvallist), allow_blank=True, showErrorMessage = False)
                ws.add_data_validation(dv)
                dv.add(d)
                ws['I{}'.format(line)] = "."
            elif dvals:
                d.value = dvals[0]
            else:
                d.value = 'huh, no desc?'

            d.alignment=Alignment(horizontal='left', vertical='center', shrink_to_fit=True)
            c = ws['K{}'.format(line)]
            c.value = mylist[0]['delcode']
            c.alignment=Alignment(horizontal='center', vertical='center', shrink_to_fit=True)
            i = ws['L{}'.format(line)]
            i.value = mylist[0]['itinitials'] # 'BY'
            i.alignment=Alignment(horizontal='center', vertical='center', shrink_to_fit=True)
            d = ws['M{}'.format(line)]
            d.value = mylist[0]['dateitcleared'].strftime("%m-%d-%y")
            d.alignment=Alignment(horizontal='center', vertical='center', shrink_to_fit=True)
            line += 1
        ws.title = "{}-{}".format(firsttag, lasttag)

        # return render_template('inventory/index.html')
        # return render_template('inventory/index.html',rows = dvals)
        

        return send_file(
                        io.BytesIO(save_virtual_workbook(wb)),
                        attachment_filename='{}Delete{}.xlsx'.format(datetime.today().strftime('%Y%m%d'), id),
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
    elif re.match("^(Edit)", submit):
        form = []
        myform = ""
        mylist = []

        # put the header fields here with a submit of "Update header detail"
        ans = _getheaderinfo(dbf, id)
        if ans:
            headerinfo = ans[0]
        else:
            flash("failed to retrieve form {}, call Bryan, and/or quit hacking.".format(id))
            return redirect(url_for('.deletions'))
        
        form.append('<input type="hidden" name="delformid" value="{}"><br />'.format(id))
        form.append('<span style="text-align:right;"><input type="submit" name="submit" value="Delete entire form {}" style="color:red;" ><br />'.format(id))
        form.append('<input type="submit" name="submit" value="Flag inactive - form {}" style="color:yellow; background-color:darkgrey;" ><br /></span>'.format(id))

        # v = "" if not 'notes' in f else f['notes']
        # i = headerinfo['id']
        myform = '<div style="border-radius: 10px; border:2px solid black; background: #cce7ff; padding:10px;">Form ID: {}<br />'.format(id)

        v = headerinfo['notes']
        v = "" if v == None else v
        # return render_template('inventory/index.html',rows = ['472', v, id, headerinfo, headerinfo['notes'], headerinfo['notes'] == None, 'blah' if headerinfo['notes'] else 'oof'])
        myform += '\n<label for="notes">Personal notes </label><input type="text" id="notes" name="notes" value="{}" ><br />'.format(v)

        v = headerinfo['schooldeleting']
        v = "" if v == None else v
        myform += '\n<label for="sdel">School Unit Deleteing Items: </label><input type="text" id="sdel" name="sdel" value="{}"><br />'.format( v)

        v = headerinfo['workorder']
        v = "" if v == None else v
        myform += '\n<label for="wo">Work Order Num: </label><input type="text" id="wo" name="wo" value="{}"><br />'.format(v)

        v = headerinfo['username']
        v = "" if v == None else v
        myform += '\n<label for="creator">Owner username: </label><input type="text" id="creator" name="creator" value="{}" required><br />'.format(v)

        myform += '\n<input type="submit" name="submit" value="Update header {}"><br />'.format(id)
        myform += '\n</div>\n&nbsp;\n'
        
        form.append(myform)

        fulldetails = _getdetails(dbf, id)

        # return render_template('inventory/index.html',rows=[{'521fulldetails':fulldetails,'v':v, 'id':id}])
        tags =  sorted(list(set([t['tag'] for t in fulldetails])))
        

        lentags = len(tags)
        # return render_template('inventory/index.html',rows = ['blue', id, tags, fulldetails])

        for tag in tags:
            mylist = []
            # form.append("I was here once, with tag{}".format(tag))
            myform = '<div style="border-radius: 10px; border:2px solid black; background: AliceBlue; padding:10px;">'

            # gather tag description(s) into mylist and work with each of those as clump
            mylist = [item for item in fulldetails if item["tag"] == tag]
            # return render_template('inventory/index.html',rows = ['blue', tag, mylist, fulldetails])

            for row in mylist:
                if myform[-7] != '10px;">':
                    myform += "<br />"
                myform += 'Row {}\n<input type="text" value="{}" name="tag{}" size="6">'.format(row['id'], row['tag'], row['id'])
                myform += '\n<input type="text" value="{}" name="description{}" >'.format(row['description'], row['id'])
                myform += '\n<input type="text" value="{}" name="delcode{}" size="3">'.format(row['delcode'], row['id'])
                myform += '\n<input type="text" value="{}" name="itinitials{}" size="3">'.format(row['itinitials'], row['id'])
                myform += '\n<input type="text" value="{}" name="dateitcleared{}" size="15">'.format(row['dateitcleared'], row['id'])
                myform += '\n<input type="submit" name="submit" value="Update detail {}">'.format(row['id'])
                myform += '\n<input type="submit" name="submit" style="color:red;" value="Delete Detail {}">'.format(row['id'])
            form.append(myform)
            form.append('\n</div>\n&nbsp;\n')

        form.append('<br /><div style="border-radius: 10px; border:2px solid black; padding:10px;">')
        form.append('Add a New Row <input type="text" value="" name="newtag" size="6">')
        form.append('<input type="text" value="" name="newdescription">')
        form.append('<input type="text" value="O" name="newdelcode" size="3">')
        form.append('<input type="text" value="{}" name="newitinitials" size="3">'.format(initials))
        form.append('<input type="text" value="{}" name="newdateitcleared" size="15">'.format(datetime.today().strftime('%m-%d-%Y')))
        form.append('<input type="submit" name="submit" value="Add New Detail line to Form {}">\n</div><br />\n'.format(id))
       
        form.append('\n<input type="submit" name="submit" value="Smartphone scan to {}"> '.format(id))
        form.append('\n<input type="submit" name="submit" value="Barcode scan or 10key to {}"> '.format(id))
        form.append('<input type="submit" name="submit" value="Download {}">'.format(id))

        return render_template('inventory/index.html', form=form, formtitle = "Edit Deletion Form {} ({} tags)".format(id, lentags))
    elif re.match("^(Cancel)", submit):
        return redirect(url_for('.deletions'))
    elif re.match("^(Smartphone scan to)", submit):
        return render_template('inventory/scan.html', pgtitle = 'scan deletions', id=id)
    elif re.match("^(Barcode scan or 10key to)", submit):
        myform = '<textarea name="scanned" rows="10" cols="20"></textarea>'
        myform += '<input type="hidden" name="id" value="{}"><br />'.format(id)
        myform += '<input type="submit" name="submit" value="Submit Tags"></<br />'
        return render_template('inventory/index.html', pgtitle = 'barcode scan deletions', form=[myform])
        
    return render_template('inventory/index.html',rows=['551catchall', id, {'submit':submit}])


@bp.route('/test', methods=["GET", "POST"])
@login_required
def test():
    # some nice form reference values: https://stackoverflow.com/questions/10434599/get-the-data-received-in-a-flask-request
    id = request.form.get('id')
    scanned = request.form.getlist("scanned")

    if id:
        pass
        # delform = DelForm(request.form.get('id'))
    if scanned:
        return render_template('inventory/index.html',rows=[scanned, request.form] )

    # cancel = request.form.get('cancel') # if key might not exist
    # if cancel:
    #     return render_template('inventory/index.html',rows=['Cancelled'] )
    # error = None

    return render_template('inventory/scan.html', pgtitle = 'blah')

@bp.route('/etest-download', methods=["GET", "POST"])
@login_required
def etest():
    # really get this from the form: 
    scanned = ['44446','57368','57360','57355','57843','57446','58818','57476','58358']
    lookedup = _getdetailsfromtags(scanned)
    # order by tag number

    # ok, this part is back to reality
    wb = load_workbook(filename = '/var/www/flaskr/flaskr/static/FARetirementBLANK.xlsx')
    # wb = load_workbook(filename = '/var/www/flaskr/flaskr/static/Book1.xlsx') 
    # wb = load_workbook(filename = '/var/www/flaskr/flaskr/static/FARetirement-dv.xlsx')

    ws = wb.active
    # ws['B25-45'] == Tag1...21
    # ws['CB25-45'] == Description
    # K: Deletion Code, O|B
    # L: IT dept, locked
    # M: date cleared, locked
    # wb.save(filename = '/var/www/flaskr/flaskr/static/FARetirementBAY.xlsx')
    line = 25
    firsttag = ""
    lasttag = ""
    # {'TagNumber':'57476','Description':'3Y TABLET REPAIR W/ADH $801-$900DOP'}
    for tag in set([tg['TagNumber'] for tg in lookedup]):
        if line == 25:
            firsttag = tag
        if (line > 45):
            ws.title = "{}-{}".format(firsttag, lasttag)
            line = 25
            firsttag = tag
            ws2 = wb.copy_worksheet(ws)
            ws = ws2
        t = ws['B{}'.format(line)]
        lasttag = tag
        t.value = tag
        t.alignment=Alignment(horizontal='center', vertical='center', shrink_to_fit=True)
        d = ws['C{}'.format(line)]
        # insert all the values here, if more than one, then make it DV, otherwise just one
        dvals = []
        for l in lookedup:
            if l['TagNumber'] == tag:
                dvals.append(l['Description'])
        if dvals and len(dvals) > 1:
            d.value = dvals[0]
            dvallist = ','.join([d.replace(',',' ') for d in dvals])
            dvallist = "Cat,Dog,Orange"
            # dv = DataValidation(type="list", formula1='"{}"'.format(dvallist), allow_blank=True)
            # # dv.prompt = 'Please select from the list'
            # # dv.promptTitle = 'List Selection'
            # # dv.error ='Your entry is not in the list'
            # # dv.errorTitle = 'Invalid Entry'
            # ws.add_data_validation(dv)
            # dv.add(d)
        elif dvals:
            d.value = dvals[0]
        else:
            d.value = 'huh, no desc?'
        # tstlst1 = ws.data_validations.dataValidation
        d.alignment=Alignment(horizontal='left', vertical='center', shrink_to_fit=True)
        c = ws['K{}'.format(line)]
        c.value = 'O' # item['DeletionCode']
        c.alignment=Alignment(horizontal='center', vertical='center', shrink_to_fit=True)
        i = ws['L{}'.format(line)]
        i.value = 'BY'
        i.alignment=Alignment(horizontal='center', vertical='center', shrink_to_fit=True)
        d = ws['M{}'.format(line)]
        d.value = '11/10/21'
        d.alignment=Alignment(horizontal='center', vertical='center', shrink_to_fit=True)
        line += 1
    ws.title = "{}-{}".format(firsttag, lasttag)

    vd = DataValidation(type="list", formula1='"3Y TABLET REPAIR W/ADH $801-$900DOP,UAG SURFACE PRO 4 CASE - BLACK,MS SURFACE PRO 4 M 128GB 4GB EDU BD Mfg#: TZ5-00001 Contract: MARKET Comes with keyboard type cas"', allow_blank=True, showErrorMessage = False)
    # class openpyxl.worksheet.datavalidation.DataValidation(type=None, formula1=None, formula2=None, showErrorMessage=True, showInputMessage=True, showDropDown=None, allowBlank=None, sqref=(), promptTitle=None, errorStyle=None, error=None, prompt=None, errorTitle=None, imeMode=None, operator=None, allow_blank=None)[source]
    # https://openpyxl.readthedocs.io/en/stable/api/openpyxl.worksheet.datavalidation.html?highlight=showErrorMessage#openpyxl.worksheet.datavalidation.DataValidation
    vd.warning = "blah"
    ws.add_data_validation(vd)
    x = ws['C25']
    x.value = '3Y TABLET REPAIR W/ADH $801-$900DOP'
    vd.error ='Your entry is not in the list'
    vd.errorTitle = 'Invalid Entry'
    vd.add(x)

    tstlst = ws.data_validations.dataValidation

    return render_template('inventory/index.html',rows = tstlst)
    # return render_template('inventory/index.html',rows = dvals)
    

    return send_file(
                    io.BytesIO(save_virtual_workbook(wb)),
                    attachment_filename='Deletes.xlsx',
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )

    # return render_template('inventory/index.html',rows=lookedup )

# example from https://stackoverflow.com/questions/11017466/flask-to-return-image-stored-in-database
@bp.route('/images/<int:pid>.jpg')
def get_image(pid):
    image_binary = read_image(pid)
    response = make_response(image_binary)
    response.headers.set('Content-Type', 'image/jpeg')
    response.headers.set(
        'Content-Disposition', 'attachment', filename='%s.jpg' % pid)
    return response

    return send_file(
        io.BytesIO(image_binary),
        mimetype='image/jpeg',
        as_attachment=True,
        attachment_filename='%s.jpg' % pid)